"""
Excel management module for storing candidate data
"""
import os
import pandas as pd
from datetime import datetime
import logging
from typing import List, Dict

# Configure logging
logging.basicConfig(level=logging.INFO)
logger = logging.getLogger(__name__)

class ExcelManager:
    """Manages Excel operations for candidate data storage"""
    
    def __init__(self, output_file: str = None):
        """Initialize Excel manager"""
        if output_file is None:
            output_file = os.path.join('data', 'candidates.xlsx')
        
        self.output_file = output_file
        self.ensure_data_directory()
    
    def ensure_data_directory(self):
        """Ensure data directory exists"""
        os.makedirs(os.path.dirname(self.output_file), exist_ok=True)
    
    def save_candidates(self, candidates: List[Dict]) -> bool:
        """Save candidate data to Excel file"""
        if not candidates:
            logger.warning("No candidates to save")
            return False
        
        try:
            # Prepare data for Excel
            excel_data = []
            
            for i, candidate in enumerate(candidates, 1):
                row_data = {
                    'ID': i,
                    'Name': candidate.get('name', 'Unknown'),
                    'Email': candidate.get('email', ''),
                    'Phone': candidate.get('phone', ''),
                    'Location': candidate.get('location', ''),
                    'Years of Experience': candidate.get('years_of_experience', 0),
                    'Experience Level': candidate.get('experience_level', 'Junior'),
                    'Applied Position': candidate.get('applied_position', ''),
                    'Skills': ', '.join(candidate.get('skills', [])),
                    'Education': candidate.get('education', ''),
                    'Overall Score': candidate.get('overall_score', 0),
                    'Contact Complete': 'Yes' if all([
                        candidate.get('email'), 
                        candidate.get('phone')
                    ]) else 'No',
                    'Extraction Source': candidate.get('extraction_source', 'email'),
                    'Processing Date': datetime.now().strftime('%Y-%m-%d %H:%M:%S')
                }
                
                # Add rating breakdown
                breakdown = candidate.get('rating_breakdown', {})
                row_data.update({
                    'Experience Points': breakdown.get('experience_points', 0),
                    'Skills Points': breakdown.get('skills_points', 0),
                    'Education Points': breakdown.get('education_points', 0),
                    'Contact Points': breakdown.get('contact_points', 0)
                })
                
                excel_data.append(row_data)
            
            # Create DataFrame
            df = pd.DataFrame(excel_data)
            
            # Save to Excel with formatting
            with pd.ExcelWriter(self.output_file, engine='openpyxl') as writer:
                # Main candidates sheet
                df.to_excel(writer, sheet_name='Candidates', index=False)
                
                # Summary sheet
                self._create_summary_sheet(writer, candidates)
                
                # Format worksheets
                self._format_worksheets(writer, df)
            
            logger.info(f"Successfully saved {len(candidates)} candidates to {self.output_file}")
            return True
            
        except Exception as e:
            logger.error(f"Error saving candidates to Excel: {str(e)}")
            return False
    
    def _create_summary_sheet(self, writer, candidates: List[Dict]):
        """Create summary statistics sheet"""
        try:
            from src.candidate_rater import CandidateRater
            rater = CandidateRater()
            summary = rater.get_rating_summary(candidates)
            
            summary_data = [
                ['Metric', 'Value'],
                ['Total Candidates', summary.get('total_candidates', 0)],
                ['Average Score', f"{summary.get('average_score', 0):.1f}"],
                [''],
                ['Experience Level Distribution', ''],
                ['Junior (0-2 years)', summary.get('level_counts', {}).get('Junior', 0)],
                ['Mid-level (2-5 years)', summary.get('level_counts', {}).get('Mid-level', 0)],
                ['Senior (5+ years)', summary.get('level_counts', {}).get('Senior', 0)],
                [''],
                ['Percentage Distribution', ''],
                ['Junior %', f"{summary.get('level_percentages', {}).get('Junior', 0):.1f}%"],
                ['Mid-level %', f"{summary.get('level_percentages', {}).get('Mid-level', 0):.1f}%"],
                ['Senior %', f"{summary.get('level_percentages', {}).get('Senior', 0):.1f}%"],
            ]
            
            summary_df = pd.DataFrame(summary_data)
            summary_df.to_excel(writer, sheet_name='Summary', index=False, header=False)
            
        except Exception as e:
            logger.error(f"Error creating summary sheet: {str(e)}")
    
    def _format_worksheets(self, writer, df):
        """Apply formatting to Excel worksheets"""
        try:
            from openpyxl.styles import Font, PatternFill, Alignment
            from openpyxl.utils.dataframe import dataframe_to_rows
            
            # Get workbook and worksheets
            workbook = writer.book
            candidates_sheet = writer.sheets['Candidates']
            
            # Header formatting
            header_font = Font(bold=True, color='FFFFFF')
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            
            # Apply header formatting
            for col in range(1, len(df.columns) + 1):
                cell = candidates_sheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center')
            
            # Auto-adjust column widths
            for column in candidates_sheet.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 2, 50)
                candidates_sheet.column_dimensions[column_letter].width = adjusted_width
            
        except Exception as e:
            logger.error(f"Error formatting worksheets: {str(e)}")
    
    def load_existing_candidates(self) -> List[Dict]:
        """Load existing candidates from Excel file"""
        if not os.path.exists(self.output_file):
            return []
        
        try:
            df = pd.read_excel(self.output_file, sheet_name='Candidates')
            candidates = df.to_dict('records')
            logger.info(f"Loaded {len(candidates)} existing candidates")
            return candidates
            
        except Exception as e:
            logger.error(f"Error loading existing candidates: {str(e)}")
            return []
    
    def append_candidates(self, new_candidates: List[Dict]) -> bool:
        """Append new candidates to existing Excel file"""
        existing_candidates = self.load_existing_candidates()
        
        # Combine with new candidates
        all_candidates = existing_candidates + new_candidates
        
        # Remove duplicates based on email
        unique_candidates = []
        seen_emails = set()
        
        for candidate in all_candidates:
            email = candidate.get('email', '').lower()
            if email and email not in seen_emails:
                unique_candidates.append(candidate)
                seen_emails.add(email)
            elif not email:  # Include candidates without email
                unique_candidates.append(candidate)
        
        return self.save_candidates(unique_candidates)
    
    def get_file_path(self) -> str:
        """Get the full path to the Excel file"""
        return os.path.abspath(self.output_file)
